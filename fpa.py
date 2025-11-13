# fpa.py — v1.4.3
# Diárias RJ — Dashboard corporativo + Tema claro/escuro + Drill-through + Metas + Export
from __future__ import annotations
import io
import os
import tempfile
from pathlib import Path
from typing import Optional, Tuple, List
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill

import plotly.express as px
import plotly.graph_objects as go

import traceback

# ========================= CONFIG BASE =========================
st.set_page_config(page_title="FPA - Diárias RJ v1.4.3", layout="wide")
st.title("📊 FPA — Diárias RJ (v1.4.3)")
st.caption("Correlação De_Para + Base_Check, tema corporativo, período, drill-through, metas e export HTML/PDF/PNG.")

ROOT = Path(".")
DIR_INPUT = ROOT / "Input"
DIR_PROC = DIR_INPUT / "Processamento"
DIR_PREM = DIR_INPUT / "Premissas"
DIR_OUT = ROOT / "output"

# ========================= SIDEBAR / TEMA ======================
with st.sidebar:
    st.header("📁 Estrutura (fixa)")
    st.write(f"**Processamento:** `{(DIR_PROC).resolve()}`")
    st.write(f"**Premissas:** `{(DIR_PREM).resolve()}`")
    st.write(f"**Saída:** `{(DIR_OUT).resolve()}`")

    sheet_base = st.text_input("Aba principal", value="Base", key="sheet_base")
    export_prefix = st.text_input(
        "Prefixo arquivo", value="BASE_Completo", key="export_prefix")
    add_ts = st.toggle("Anexar timestamp no nome", value=True, key="add_ts")

    st.divider()
    st.header("🎨 Tema")
    theme_mode = st.radio("Modo", ["Claro", "Escuro"], horizontal=True)
    # Paleta da marca — permite ajuste rápido
    st.caption("Paleta (pode ajustar as cores):")
    col_ok = st.color_picker("Convergente", "#2ca02c", key="c_ok")
    col_div = st.color_picker("Divergente", "#d62728", key="c_div")
    col_sr = st.color_picker("Sem Relação", "#7f7f7f", key="c_sr")

    st.divider()
    st.header("🎯 Meta")
    meta_conv = st.slider("% Convergência (objetivo)",
                          min_value=0, max_value=100, value=80, step=1)

# Paleta aplicada
COLOR_MAP = {"SETOR CONVERGENTE": col_ok,
             "SETOR DIVERGENTE": col_div, "**SEM RELAÇÃO**": col_sr}
PLOT_TEMPLATE = "plotly_dark" if theme_mode == "Escuro" else "simple_white"

# CSS leve
CUSTOM_CSS = f"""
<style>
html, body, [class^="css"] {{
  font-family: 'Inter', 'Segoe UI', Arial, Helvetica, sans-serif !important;
}}
.block-container {{ padding-top: 1.0rem; padding-bottom: 1.6rem; }}
.kpi {{ padding: 12px 16px; border-radius: 12px; border: 1px solid {"#333" if theme_mode == "Escuro" else "#eee"}; 
        background: {"#111" if theme_mode == "Escuro" else "#fafafa"} }}
.kpi .title {{ font-size: 0.85rem; opacity: .8; margin-bottom: 4px; }}
.kpi .value {{ font-weight: 800; font-size: 1.25rem; }}
.kpi .delta {{ font-size: 0.85rem; opacity: .7; }}
.stDownloadButton button {{ border-radius: 10px !important; }}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ========================= HELPERS =============================


def now_stamp() -> str: return datetime.now().strftime("%Y%m%d_%H%M%S")


def safe_name(
    s: str) -> str: return "".join(ch for ch in s if ch not in r'\/:*?"<>|').strip()


def norm_cols(df: pd.DataFrame) -> pd.DataFrame: df.columns = [
    str(c).strip() for c in df.columns]; return df


def norm(x):
    if pd.isna(x) or x is None:
        return None
    return str(x).strip()


def find_col(cols, options: List[str]) -> Optional[str]:
    m = {str(c).strip().lower(): c for c in cols}
    for opt in options:
        k = opt.strip().lower()
        if k in m:
            return m[k]
    return None


def insert_after(df: pd.DataFrame, after_col: str, new_col: str, values) -> pd.DataFrame:
    cols = list(df.columns)
    idx = cols.index(after_col)
    return pd.concat([df.iloc[:, :idx+1], pd.Series(values, name=new_col, index=df.index), df.iloc[:, idx+1:]], axis=1)


def load_processing_file(proc_dir: Path) -> Path:
    if not proc_dir.exists():
        raise FileNotFoundError(
            f"Pasta de processamento inexistente: {proc_dir}")
    files = sorted(proc_dir.glob("*.xlsx"),
                   key=lambda x: x.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(
            "Nenhum arquivo encontrado em Input/Processamento.")
    f = files[0]
    if not f.name.startswith("Produção_Financeira_Regional_"):
        raise ValueError("Arquivo não Fora do Padrão de Processamento!")
    return f


def load_premissas(prem_dir: Path) -> Tuple[Path, Path]:
    depara = prem_dir / "Base_De_Para_RegRJ.xlsx"
    check = prem_dir / "Base_Check.xlsx"
    if not depara.exists():
        raise FileNotFoundError(f"Premissa ausente: {depara}")
    if not check.exists():
        raise FileNotFoundError(f"Premissa ausente: {check}")
    return depara, check


def processar(df_base: pd.DataFrame, map_df: pd.DataFrame, chk_df: pd.DataFrame) -> pd.DataFrame:
    df_base = norm_cols(df_base)
    map_df = norm_cols(map_df)
    chk_df = norm_cols(chk_df)
    # De_Para
    ds_col = find_col(df_base.columns, [
                      "ds_pro_fat", "DS_PRO_FAT", "ds pro fat"])
    if ds_col is None:
        raise KeyError("Coluna 'ds_pro_fat' não encontrada na aba Base.")
    ds_map = find_col(
        map_df.columns, ["ds_pro_fat", "DS_PRO_FAT", "ds pro fat"])
    para_col = find_col(map_df.columns, ["Para", "para", "PARA"])
    if ds_map is None or para_col is None:
        raise KeyError("Premissa De_Para sem 'ds_pro_fat' e/ou 'Para'.")
    mapping = {norm(r[ds_map]): r[para_col]
               for _, r in map_df.iterrows() if norm(r[ds_map])}
    diaria = df_base[ds_col].map(norm).map(
        lambda x: mapping.get(x, "**SEM RELAÇÃO**"))
    df1 = insert_after(df_base, ds_col, "Diária Padrão", diaria)
    # Base_Check
    diaria_chk = find_col(
        chk_df.columns, ["Diária Padrão", "Diaria Padrao", "Diaria Padrão"])
    setor_chk = find_col(chk_df.columns, ["Setor Padrão", "Setor Padrao"])
    if diaria_chk is None or setor_chk is None:
        raise KeyError(
            "Premissa Base_Check sem 'Diária Padrão' e/ou 'Setor Padrão'.")
    chk_map = {norm(r[diaria_chk]): r[setor_chk]
               for _, r in chk_df.iterrows() if norm(r[diaria_chk])}
    setor_padrao = df1["Diária Padrão"].map(
        lambda x: chk_map.get(x, "**SEM RELAÇÃO**"))
    df2 = insert_after(df1, "Diária Padrão", "Setor Padrão", setor_padrao)
    # Check
    setor_orig_col = find_col(df2.columns, [
                              "ds_grupo_macro", "Setor", "setor", "Setor Original", "setor original", "DS_GRUPO_MACRO"])

    def _check(row):
        sp = norm(row["Setor Padrão"])
        if sp == "**SEM RELAÇÃO**":
            return "**SEM RELAÇÃO**"
        if setor_orig_col:
            so = norm(row[setor_orig_col])
            return "SETOR CONVERGENTE" if so == sp else "SETOR DIVERGENTE"
        return "SETOR CONVERGENTE" if (sp or "").upper() == "INTERNAÇÃO" else "SETOR DIVERGENTE"
    df2 = insert_after(df2, "Setor Padrão", "Check", df2.apply(_check, axis=1))
    return df2


def to_excel_bytes_clean(df: pd.DataFrame, sheet_name: str) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb[sheet_name]
    mr, mc = ws.max_row, ws.max_column
    empty_border = Border()
    empty_fill = PatternFill(fill_type=None)
    for row in ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc):
        for cell in row:
            cell.font = Font(name="Arial", size=9, bold=False)
            cell.border = empty_border
            cell.fill = empty_fill
    for cell in ws[1]:
        cell.font = Font(name="Arial", size=9, bold=True)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def export_html(figs: List[go.Figure], title: str, extra_html: str = "") -> bytes:
    parts = [f.to_html(full_html=False, include_plotlyjs=False) for f in figs]
    page = f"""<!doctype html><html><head><meta charset="utf-8">
<title>{title}</title><script src="https://cdn.plot.ly/plotly-2.32.0.min.js"></script>
<style>body{{font-family:Arial,Helvetica,sans-serif; margin:20px}} h1{{font-size:20px}}</style>
</head><body><h1>{title}</h1>{extra_html}{''.join(parts)}</body></html>"""
    return page.encode("utf-8")


def try_export_pdf(html_bytes: bytes) -> Optional[bytes]:
    try:
        import pdfkit
        with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
            f.write(html_bytes)
            f.flush()
            hp = f.name
        pdf_bytes = pdfkit.from_file(hp, False)
        os.remove(hp)
        return pdf_bytes
    except Exception:
        pass
    try:
        from weasyprint import HTML
        return HTML(string=html_bytes.decode("utf-8")).write_pdf()
    except Exception:
        return None


# ========================= PIPELINE ============================
try:
    # Carregar insumos
    proc_file = load_processing_file(DIR_PROC)
    path_depara, path_check = load_premissas(DIR_PREM)

    df_base = pd.read_excel(proc_file, sheet_name=sheet_base, dtype=object)
    map_df = pd.read_excel(path_depara, dtype=object)
    chk_df = pd.read_excel(path_check, dtype=object)

    df_out = processar(df_base, map_df, chk_df)

    # ===== Cabeçalho / logo opcional =====
    logo_path = DIR_PREM / "logo.png"
    if logo_path.exists():
        c1, c2 = st.columns([1, 4])
        with c1:
            st.image(str(logo_path))
        with c2:
            st.markdown("### Painel — Produção de Diárias")
            st.caption(proc_file.name)

    # ===== Detectar colunas =====
    col_valor = find_col(
        df_out.columns, ["PRODr", "PROD", "Valor", "valor", "vr_prod", "vr_total"])
    if col_valor is None:
        raise KeyError("Coluna de valores não encontrada (ex.: 'PRODr').")
    col_setor_macro = find_col(
        df_out.columns, ["ds_grupo_macro", "DS_GRUPO_MACRO", "Setor", "setor"])
    if col_setor_macro is None:
        raise KeyError(
            "Coluna de agrupamento setorial não encontrada (ex.: 'ds_grupo_macro').")
    col_diaria = "Diária Padrão"
    col_check = "Check"

    # ===== Período (opcional) =====
    col_data = find_col(df_out.columns, [
                        "data", "dt", "Data", "DT", "competencia", "Competência", "PERIODO", "Período", "periodo"])
    dfv = df_out.copy()
    if col_data:
        dfv["_data"] = pd.to_datetime(df_out[col_data], errors="coerce")
        mind, maxd = pd.to_datetime(
            dfv["_data"].min()), pd.to_datetime(dfv["_data"].max())
        with st.sidebar:
            st.divider()
            st.header("🗓️ Período")
            d_ini, d_fim = st.date_input(
                "Intervalo", value=(mind.date(), maxd.date()))
        dfv = dfv[(dfv["_data"].dt.date >= d_ini) &
                  (dfv["_data"].dt.date <= d_fim)]

    # ===== Tabela virtual =====
    st.subheader("📄 Tabela Base — processada")
    st.dataframe(dfv, use_container_width=True, hide_index=True)

    # ===== Drill-down filters =====
    st.markdown("---")
    st.subheader("🎯 Drill-down")
    f1, f2, f3 = st.columns(3)
    with f1:
        opt_setor = [
            "(Todos)"] + sorted(dfv["Setor Padrão"].dropna().astype(str).unique().tolist())
        sel_setor = st.selectbox("Setor Padrão", opt_setor, index=0)
    df_d1 = dfv if sel_setor == "(Todos)" else dfv[dfv["Setor Padrão"].astype(
        str) == sel_setor]

    with f2:
        opt_diaria = [
            "(Todos)"] + sorted(df_d1["Diária Padrão"].dropna().astype(str).unique().tolist())
        sel_diaria = st.selectbox("Diária Padrão", opt_diaria, index=0)
    df_d2 = df_d1 if sel_diaria == "(Todos)" else df_d1[df_d1["Diária Padrão"].astype(
        str) == sel_diaria]

    with f3:
        if "ds_pro_fat" in df_d2.columns:
            opt_ds = [
                "(Todos)"] + sorted(df_d2["ds_pro_fat"].dropna().astype(str).unique().tolist())
        else:
            opt_ds = ["(N/D)"]
        sel_ds = st.selectbox("ds_pro_fat", opt_ds, index=0)
    df_d3 = df_d2 if (sel_ds in ["(Todos)", "(N/D)"]
                      ) else df_d2[df_d2["ds_pro_fat"].astype(str) == sel_ds]

    # ===== KPIs =====
    df_dash = df_d3.copy()
    df_dash["_valor"] = pd.to_numeric(
        df_dash[col_valor], errors="coerce").fillna(0)
    tot = float(df_dash["_valor"].sum())
    conv = float(df_dash.loc[df_dash[col_check] ==
                 "SETOR CONVERGENTE", "_valor"].sum())
    div = float(df_dash.loc[df_dash[col_check] ==
                "SETOR DIVERGENTE", "_valor"].sum())
    sr = float(df_dash.loc[(df_dash[col_diaria] == "**SEM RELAÇÃO**")
               | (df_dash["Setor Padrão"] == "**SEM RELAÇÃO**"), "_valor"].sum())

    def pct(x): return 0 if tot == 0 else x/tot*100
    pct_conv = pct(conv)

    kc1, kc2, kc3, kc4 = st.columns(4)
    kc1.markdown(f"<div class='kpi'><div class='title'>Total (∑ PRODr)</div><div class='value'>{tot:,.2f}".replace(
        ",", "X").replace(".", ",").replace("X", ".") + "</div></div>", unsafe_allow_html=True)
    kc2.markdown(f"<div class='kpi'><div class='title'>Convergente</div><div class='value'>{conv:,.2f}".replace(",", "X").replace(
        ".", ",").replace("X", ".") + f"</div><div class='delta'>{pct_conv:.1f}%</div></div>", unsafe_allow_html=True)
    kc3.markdown(f"<div class='kpi'><div class='title'>Divergente</div><div class='value'>{div:,.2f}".replace(",", "X").replace(
        ".", ",").replace("X", ".") + f"</div><div class='delta'>{pct(div):.1f}%</div></div>", unsafe_allow_html=True)
    status = "✅ Atingiu meta" if pct_conv >= meta_conv else "⚠️ Abaixo da meta"
    kc4.markdown(
        f"<div class='kpi'><div class='title'>Meta Convergência</div><div class='value'>{meta_conv:.0f}%</div><div class='delta'>{status}</div></div>", unsafe_allow_html=True)

    # ===== Agregações p/ gráficos =====
    grp = (df_dash.groupby([col_setor_macro, col_diaria, col_check], dropna=False)[
           "_valor"].sum().reset_index())

    # ===== Gráfico 1 (Setor) — drill-through por clique =====
    st.markdown("#### 🔹 Por Setor (empilhado por Check)")
    df_setor = grp.groupby([col_setor_macro, col_check])[
        "_valor"].sum().reset_index()
    g1 = px.bar(df_setor, x="_valor", y=col_setor_macro, color=col_check, orientation="h",
                barmode="stack", hover_data={"_valor": ":,.2f"}, color_discrete_map=COLOR_MAP, template=PLOT_TEMPLATE)
    g1.update_layout(margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
    click1 = st.plotly_chart(g1, use_container_width=True,
                             key="g1", on_select="ignore")

    # ===== Gráfico 2 (Diária) — drill-through por clique =====
    st.markdown("#### 🔹 Por Diária Padrão (empilhado por Check)")
    df_diaria = grp.groupby([col_diaria, col_check])[
        "_valor"].sum().reset_index()
    g2 = px.bar(df_diaria, x=col_diaria, y="_valor", color=col_check, barmode="stack",
                hover_data={"_valor": ":,.2f"}, color_discrete_map=COLOR_MAP, template=PLOT_TEMPLATE)
    g2.update_layout(margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
    click2 = st.plotly_chart(g2, use_container_width=True,
                             key="g2", on_select="ignore")

    # ===== Gráfico 3 (Pizza) =====
    st.markdown(
        "#### 🔹 Distribuição Geral (Convergente × Divergente × Sem Relação)")
    df_pie = (df_dash
              .assign(_grupo=lambda d: d[col_check].where(~((d[col_diaria] == "**SEM RELAÇÃO**")
                                                            | (d["Setor Padrão"] == "**SEM RELAÇÃO**")),
                                                          "**SEM RELAÇÃO**"))
              .groupby("_grupo")["_valor"].sum().reset_index().rename(columns={"_grupo": "Grupo"}))
    g3 = px.pie(df_pie, names="Grupo", values="_valor", hole=0.35,
                color="Grupo", color_discrete_map=COLOR_MAP, template=PLOT_TEMPLATE)
    g3.update_traces(textposition='inside',
                     texttemplate='%{label}<br>%{percent:.1%}')
    g3.update_layout(margin=dict(l=10, r=10, t=10, b=10))
    st.plotly_chart(g3, use_container_width=True)

    # ===== Drill-through detail (usa seleção manual pelos filtros clicados)
    st.markdown("#### 🧩 Detalhe do Drill-through")
    # Recupera última seleção do usuário (via estado manual)
    if "sel_setor_from_chart" not in st.session_state:
        st.session_state.sel_setor_from_chart = None
    if "sel_diaria_from_chart" not in st.session_state:
        st.session_state.sel_diaria_from_chart = None

    # Em Streamlit atual, pegamos o clique via session_state do componente (workaround: instruir uso dos filtros acima).
    # Para facilitar, mostramos a mesma visão filtrável e um botão de limpar:
    df_detail = df_dash.copy()
    if sel_setor != "(Todos)":
        df_detail = df_detail[df_detail["Setor Padrão"].astype(
            str) == sel_setor]
    if sel_diaria != "(Todos)":
        df_detail = df_detail[df_detail["Diária Padrão"].astype(
            str) == sel_diaria]
    if sel_ds not in ["(Todos)", "(N/D)"] and "ds_pro_fat" in df_detail.columns:
        df_detail = df_detail[df_detail["ds_pro_fat"].astype(str) == sel_ds]

    st.dataframe(df_detail, use_container_width=True, hide_index=True)
    if st.button("Limpar seleção (drill)"):
        st.rerun()

    # ===== Exportações =====
    st.markdown("---")
    st.subheader("💾 Export")
    xlsb = to_excel_bytes_clean(df_out, sheet_name=sheet_base)
    fname = f"{safe_name(export_prefix)}_{now_stamp() if add_ts else ''}.xlsx".replace(
        '__', '.')
    st.download_button("Excel (Base processada)", data=xlsb,
                       file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # HTML / PDF do dashboard
    html_bytes = export_html([g1, g2, g3], title="Painel — Diárias RJ",
                             extra_html=f"<p><strong>Arquivo:</strong> {proc_file.name}</p>")
    st.download_button("🌐 Dashboard (HTML)", data=html_bytes,
                       file_name="dashboard_diarias.html", mime="text/html")
    pdf_bytes = try_export_pdf(html_bytes)
    if pdf_bytes:
        st.download_button("🧾 Dashboard (PDF)", data=pdf_bytes,
                           file_name="dashboard_diarias.pdf", mime="application/pdf")
    else:
        st.info("PDF opcional: instale `wkhtmltopdf` (pdfkit) ou os requisitos do `weasyprint`. O HTML já está disponível.")

    # PNGs (kaleido)
    try:
        png1 = g1.to_image(format="png", scale=2)
        png2 = g2.to_image(format="png", scale=2)
        png3 = g3.to_image(format="png", scale=2)
        c1, c2, c3 = st.columns(3)
        with c1:
            st.download_button("⬇️ Setor (PNG)", data=png1,
                               file_name="grafico_setor.png", mime="image/png")
        with c2:
            st.download_button("⬇️ Diária (PNG)", data=png2,
                               file_name="grafico_diaria.png", mime="image/png")
        with c3:
            st.download_button("⬇️ Distribuição (PNG)", data=png3,
                               file_name="grafico_distribuicao.png", mime="image/png")
    except Exception:
        st.warning("Para exportar PNG, instale `kaleido`.")

    # salva Excel no /output
    DIR_OUT.mkdir(parents=True, exist_ok=True)
    with open(DIR_OUT / fname, "wb") as f:
        f.write(xlsb)
    st.success(f"📦 Excel salvo em: {(DIR_OUT / fname).resolve()}")

except Exception as e:
    st.error("❌ Erro inesperado ao rodar o painel.")
    st.code(traceback.format_exc())
    st.stop()
